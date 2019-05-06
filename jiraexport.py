#
# skipped Pandas and using tradional excel reading lib: import openpyxl 
#


from jira import JIRA
from datetime import datetime
import logging as log
#import pandas 
import argparse
import getpass
import time
import sys, logging
from author import Authenticate  # no need to use as external command
from author import DoJIRAStuff
import openpyxl 
from collections import defaultdict
import re

start = time.clock()
__version__ = u"0.1.RISKS" 

# should pass via parameters
ENV="DEV"
#ENV=u"PROD"

logging.basicConfig(level=logging.DEBUG) # IF calling from Groovy, this must be set logging level DEBUG in Groovy side order these to be written out


def main():

    
    JIRASERVICE=u""
    JIRAPROJECT=u""
    PSWD=u''
    USER=u''
  
    logging.debug (u"--Python starting checking Jira issues for attachemnt adding --") 

 
    parser = argparse.ArgumentParser(usage="""
    {1}    Version:{0}     -  mika.nokka1@gmail.com 
    
    USAGE:
    -filepath  | -p <Path to Excel file directory>
    -filename   | -n <Excel filename>

    """.format(__version__,sys.argv[0]))

    #parser.add_argument('-f','--filepath', help='<Path to attachment directory>')
    parser.add_argument('-q','--excelfilepath', help='<Path to excel directory>')
    parser.add_argument('-n','--filename', help='<Excel filename>')
    
    parser.add_argument('-v','--version', help='<Version>', action='store_true')
    
    parser.add_argument('-w','--password', help='<JIRA password>')
    parser.add_argument('-u','--user', help='<JIRA user>')
    parser.add_argument('-s','--service', help='<JIRA service>')
    parser.add_argument('-l','--links', help='<Target Jira project to which issues to be linked>') #add issue links to generated issues (target "into" linked issues must be allready in target jira)
    parser.add_argument('-p','--project', help='<Target JIRA project to be created>')
    #parser.add_argument('-z','--rename', help='<rename files>') #adhoc operation activation
    #parser.add_argument('-x','--ascii', help='<ascii file names>') #adhoc operation activation
        
    args = parser.parse_args()
    
    if args.version:
        print 'Tool version: %s'  % __version__
        sys.exit(2)    
           
    #filepath = args.filepath or ''
    excelfilepath = args.excelfilepath or ''
    filename = args.filename or ''
    
    JIRASERVICE = args.service or ''
    JIRAPROJECT = args.project or ''
    PSWD= args.password or ''
    USER= args.user or ''
    LINKS=args.links or ''
    #RENAME= args.rename or ''
    #ASCII=args.ascii or ''
    
    # quick old-school way to check needed parameters
    if (JIRASERVICE=='' or PSWD=='' or USER==''  or excelfilepath=='' or JIRAPROJECT=='' or filename==''):
        parser.print_help()
        print "args: {0}".format(args)
        sys.exit(2)

    
    
    Authenticate(JIRASERVICE,PSWD,USER)
    jira=DoJIRAStuff(USER,PSWD,JIRASERVICE)
    
    excel=excelfilepath+"/"+filename
    logging.debug ("Excel file:{0}".format(excel))

    Issues=defaultdict(dict) 
    MainSheet="Sheet0" 
    wb= openpyxl.load_workbook(excel)
    #types=type(wb)
    #logging.debug ("Type:{0}".format(types))
    #sheets=wb.get_sheet_names()
    #logging.debug ("Sheets:{0}".format(sheets))
    CurrentSheet=wb[MainSheet] 
    
    
     ########################################
    #CONFIGURATIONS AND EXCEL COLUMN MAPPINGS, both main and subtask excel
    DATASTARTSROW=4 # data section starting line 
    A=1 #issuetype
    E=5 #SUMMARY
    F=6 #priority
    H=8 #Status   
    Q=17 #Assignee
    S=19 #Disciopline(F)
    
    T=20 #Probability
    U=21 #HSE Impact
    V=22 #Schedule Impact 
    W=23 #Quality Impact
    
    Z=26 #Risk Cost
    AK=37 #Linked Issues
    AM=39 ##Disciopline(RM)
    AN=40 #Description
    AB=28 #Mitigation Costs (Keur)
    
 
    ##############################################################################################
    #Go through main excel sheet for main issue keys (and contents findings)
    # Create dictionary structure (remarks)
    # NOTE: Uses hardcoded sheet/column values
    # NOTE: As this handles first sheet, using used row/cell reading (buggy, works only for first sheet) 
    #
    i=DATASTARTSROW # brute force row indexing
    for row in CurrentSheet[('C{}:C{}'.format(DATASTARTSROW,CurrentSheet.max_row))]:  # go trough all column C  rows (issue key when imp exp eported)
        for mycell in row:
            KEY=mycell.value
            logging.debug("ROW:{0} Original ID:{1}".format(i,mycell.value))
            Issues[KEY]={} # add to dictionary as master key (KEY)
            
            #Just hardocode operations, POC is one off
            #LINKED_ISSUES=(CurrentSheet.cell(row=i, column=K).value) #NOTE THIS APPROACH GOES ALWAYS TO THE FIRST SHEET
            #logging.debug("Attachment:{0}".format((CurrentSheet.cell(row=i, column=K).value))) # for the same row, show also column K (LINKED_ISSUES) values
            #Issues[KEY]["LINKED_ISSUES"] = LINKED_ISSUES
            
            SUMMARY=(CurrentSheet.cell(row=i, column=E).value)
            if not SUMMARY:
                SUMMARY="Summary for this task has not been defined"
            Issues[KEY]["SUMMARY"] = SUMMARY
            
            ISSUE_TYPE=(CurrentSheet.cell(row=i, column=A).value)
            Issues[KEY]["ISSUE_TYPE"] = ISSUE_TYPE
            
            STATUS=(CurrentSheet.cell(row=i, column=E).value)
            Issues[KEY]["SUMMARY"] = SUMMARY
            
            PRIORITY=(CurrentSheet.cell(row=i, column=F).value)
            Issues[KEY]["PRIORITY"] = PRIORITY
            
            STATUS=(CurrentSheet.cell(row=i, column=H).value)
            Issues[KEY]["STATUS"] = STATUS
            
            
            ASSIGNEE=(CurrentSheet.cell(row=i, column=Q).value)
            Issues[KEY]["ASSIGNEE"] = ASSIGNEE
            
            DisciplineF=(CurrentSheet.cell(row=i, column=S).value)
            Issues[KEY]["DisciplineF"] = DisciplineF
            
            DisciplineRM=(CurrentSheet.cell(row=i, column=AM).value)
            Issues[KEY]["DisciplineRM"] = DisciplineRM
            
            DESCRIPTION=(CurrentSheet.cell(row=i, column=AN).value)
            Issues[KEY]["DESCRIPTION"] = DESCRIPTION
            
            PROBABILITY=(CurrentSheet.cell(row=i, column=T).value)
            Issues[KEY]["PROBABILITY"] = PROBABILITY
            
            HSEImpact=(CurrentSheet.cell(row=i, column=U).value)
            Issues[KEY]["HSEImpact"] = HSEImpact
            
              
            #RESPHONE=(CurrentSheet.cell(row=i, column=U).value)
            #Issues[KEY]["RESPHONE"] = RESPHONE
            
            SheduleImpact=(CurrentSheet.cell(row=i, column=V).value)
            Issues[KEY]["SheduleImpact"] = SheduleImpact
            
            QualityImpact=(CurrentSheet.cell(row=i, column=W).value)
            Issues[KEY]["QualityImpact"] = QualityImpact
            
            RiskCost=(CurrentSheet.cell(row=i, column=Z).value)
            Issues[KEY]["RiskCost"] = RiskCost
        
            MitigationCostsKeur=(CurrentSheet.cell(row=i, column=AB).value)
            Issues[KEY]["MitigationCostsKeur"] = MitigationCostsKeur
        
                
            LinkedIssues=(CurrentSheet.cell(row=i, column=AK).value)
            Issues[KEY]["LinkedIssues"] = LinkedIssues
            
            

            logging.debug("---------------------------------------------------")
            i=i+1
    
    #print Issues.items() 
    
    
    for key, value in Issues.iteritems() :
        print "++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        print "KEY: {0}".format(key)
        #print key, value
        
        # check linked issues values form each issue found from excel
        one=Issues.get(key)
        for key, value in one.iteritems() :
            print "************************************************************************"
            if (value==None):
                if (isinstance(value, long)):
                    castedValue=0
                    print "NONE NUMBER"
                else:
                    castedValue=0
                    print "NONE" 
            else:
                if (isinstance(value, long)): # is it number??
                    castedValue=value # numbers dont need utf-8 endocing
                else:
                    castedValue=value.encode('utf-8')   
            
            print "{0} {1}".format(key,castedValue)
            
            
            if (LINKS and key=="LinkedIssues"): #-l parameter to do links operations to given target project
                #print "Linked issues column found"
                
                print "Linking active: Linking target project: {0}".format(LINKS)
                if (value==None): #no linked items case
                    value2="NONE"
                    print "No linked issues found"
                else:
                    value2=value
                onelink=value2.split(':')
                for item in onelink :
                    #print "value:{0}".format(item.encode('utf-8'))
                    regex = r"(.*)(')(.*)(')"   #TT1400-39 'Logistic plan to do' (Risk Mitigation)
                    match = re.search(regex, item)
                
                    if (match):
                        hit=match.group(3)
                        #print "-----------------------------------------------------------"
                        print "Linked issue Summmary ==>  {0}".format(hit.encode('utf-8'))
                        #print "-----------------------------------------------------------"
                        
                        #project = "Risk Mitigation Panel Line"  and summary ~ "Kuitulaser hankinta ja hitsauslaboratorion hankinta"
                        #issue_list = jira.search_issues("Project = {0} and Summary ~ {1}".format(LINKS,hit))
                        
                        jql_query="Project = \'{0}\' and Summary ~ \'{1}\'".format(LINKS,hit.encode('utf-8'))
                        #print "Query:{0}".format(jql_query)
                        
                        issue_list=jira.search_issues(jql_query)
                        
                        if len(issue_list) == 1:
                            for issue in issue_list:
                                #logging.debug("One issue returned for query")
                                logging.debug("ISSUE TO BE LINKED ==> {0}".format(issue))
                
                        elif len(issue_list) > 1:
                            logging.debug("ERROR ==> More than 1 issue was returned by JQL query")
                        else:
                            logging.debug("==> No issue(s) returned by JQL query")
                            
                        time.sleep(0.7)
                
            if (key=="ASSIGNEE"):
                print "Assignee column found"
                
                if (value==None): #no linked items case
                    value2="NONE"
                    print "No assignee found"
                    USERNAME_ASSIGNEE="-1"
                else:
                    value2=value
                    regex = r"(.*)(\()(.*)(\))"   #Korpela, Matias (korpma11)
                    match = re.search(regex, value2)
                
                    if (match):
                        USERNAME_ASSIGNEE=match.group(3).encode('utf-8')
                        #print "-----------------------------------------------------------"
                        print "Assignee username ==>  {0}".format(USERNAME_ASSIGNEE)
                        #print "-----------------------------------------------------------"
                    else:
                        USERNAME_ASSIGNEE="-1"
                        
                
            if (key=="MitigationCostsKeur"):
                print "Mitigation cost column found"
                MitigationCostsKeur=castedValue
                       
                        
                        
            
        CreateMitigationIssue(jira,JIRAPROJECT,SUMMARY,ISSUE_TYPE,PRIORITY,STATUS,USERNAME_ASSIGNEE,DESCRIPTION,MitigationCostsKeur)
        sys.exit(5) #testinf do once
        print "++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
    #now excel has been prosessed
        
    # test all with one item
    TESTING=0
    if (TESTING):
        print "TESTTESTETESTTESTETE ------------------------------ TESTETRSTESTS"
        print "INV91649RM-18"
        print Issues.get("INV91649RM-18")
        
        one=Issues.get("INV91649RM-18")
        for key, value in one.iteritems() :
            print "************************************************************************"
            if (value==None):
                    castedValue=""
                    print "NONE"
            else:
                    castedValue=value.encode('utf-8')
            print "{0} {1}".format(key,castedValue)
            print
            if (key=="LinkedIssues"):
                print "Linked issues found"
                onelink=value.split(':')
                for item in onelink :
                    print "value:{0}".format(item)
                    regex = r"(.*)(')(.*)(')"   #TT1400-39 'Logistic plan to do' (Risk Mitigation)
                    match = re.search(regex, item)
                    
                    if (match):
                        hit=match.group(3)
                        print "-----------------------------------------------------------"
                        print "Linked issue Summmary: {0}".format(hit)
                        print "-----------------------------------------------------------"
                                        
                        #project = "Risk Mitigation Panel Line"  and summary ~ "Kuitulaser hankinta ja hitsauslaboratorion hankinta"
                        issue_list = jira.search_issues("Project = {0} and Summary ~ {1}".format(LINKS,hit))
                        if len(issue_list) == 1:
                            for issue in issue_list:
                                logging.debug("One issue returned for query")
                
                        elif len(issue_list) > 1:
                            logging.debug("More than 1 issue was returned by JQL query")
                            
                
                        else:
                            logging.debug("No issue(s) returned by JQL query")
                            
                        time.sleep(0.7)
                    
    
    end = time.clock()
    totaltime=end-start
    print "Time taken:{0} seconds".format(totaltime)
       
            
    print "*************************************************************************"
    

       
    sys.exit(0)
    
    
def CreateMitigationIssue(jira,JIRAPROJECT,SUMMARY,ISSUE_TYPE,PRIORITY,STATUS,USERNAME_ASSIGNEE,DESCRIPTION,MitigationCostsKeur):
    
    
    
    jiraobj=jira
    project=JIRAPROJECT
    TASKTYPE="Task" #hardcoded

    #'resolution': STATUS,
    #dev low =10002, high=10000, medium=10001
    
    print "Creating mitigation issue for JIRA project: {0}".format(project)
    
    
    issue_dict = {
    'project': {'key': JIRAPROJECT},
    'summary': SUMMARY,
    'description': DESCRIPTION,
    'issuetype': {'name': TASKTYPE},
    'priority': {'name': PRIORITY }, 
    #'resolution':{'id': '10100'},
    'assignee': {'name':USERNAME_ASSIGNEE},
    
    'customfield_14302' if (ENV =="DEV") else 'customfield_14216' : int(MitigationCostsKeur),  # MitigationCostsKeur dev: 14302  prod: 14216


    }

    try:
        new_issue = jiraobj.create_issue(fields=issue_dict)
        print "Issue created OK"
        
        
    except Exception,e:
        print("Failed to create JIRA object, error: %s" % e)
        sys.exit(1)
    return new_issue    
    
     
 
    
    
if __name__ == '__main__':
    main()
    
    
    
    

    
    
    