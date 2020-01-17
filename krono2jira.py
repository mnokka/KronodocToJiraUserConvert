#!/usr/bin/python
# -*- coding: latin-1 -*-

# Uses excel with Kronodoc exported Jira data
# Changes perosn name to Jira useraccount (or if not exists , uses default jira account)
# Idea is to make excel Jira import to possible
# 


import os, sys


from jira import JIRA
from datetime import datetime
import logging as log
#import pandas 
import argparse
import getpass
import time
import sys, logging
from author import Authenticate  # no need to use as external command
from author import DoJIRAStuff
import openpyxl 
from collections import defaultdict
import re
from pprint import pprint
from dns.name import empty

start = time.clock()
__version__ = u"0.9.KRONODOC" 



###################################################################
# TODO: should pass via parameters
# CODE CONFIGURATIONS
#####################################################################

logging.basicConfig(level=logging.DEBUG) # IF calling from Groovy, this must be set logging level DEBUG in Groovy side order these to be written out

# development vs production Jira
#ENV="DEV"
ENV="PROD"


# do only one operation for testing purposes
ONCE="NO"
#ONCE="YES"


##################################################
#CONFIGURATIONS AND EXCEL COLUMN MAPPINGS

DATASTARTSROW=2  # 2 # data section starting line 
J=10 # this column holds the kronodoc exporter usernames , CHECK ALSO CODE FOR COLUMN USAGE
MainSheet="CCL2"    # HARDCODED SHEET 


####END OF CONFIGURATIONS #######################################################################








def main():

    
    JIRASERVICE=u""
    JIRAPROJECT=u""
    PSWD=u''
    USER=u''
  
    logging.debug (u"--Python starting checking Jira issues for attachemnt adding --") 

 
    parser = argparse.ArgumentParser(usage="""
    {1}    Version:{0}     -  mika.nokka1@gmail.com 
    


    """.format(__version__,sys.argv[0]))


    parser.add_argument('-q','--excelfilepath', help='<Path to excel directory>')
    parser.add_argument('-n','--filename', help='<Issues excel filename>')
    
    parser.add_argument('-v','--version', help='<Version>', action='store_true')
    
    parser.add_argument('-w','--password', help='<JIRA password>')
    parser.add_argument('-u','--user', help='<JIRA username>')
    parser.add_argument('-s','--service', help='<JIRA service, like https://my.jira.com>')
    

        
    args = parser.parse_args()
    
    if args.version:
        print 'Tool version: %s'  % __version__
        sys.exit(2)    
           
    #filepath = args.filepath or ''
    excelfilepath = args.excelfilepath or ''
    filename = args.filename or ''
    
    JIRASERVICE = args.service or ''
    PSWD= args.password or ''
    USER= args.user or ''
    
    # quick old-school way to check needed parameters
    if (JIRASERVICE=='' or PSWD=='' or USER==''  or excelfilepath=='' or filename==''):
        parser.print_help()
        #print "args: {0}".format(args)
        sys.exit(2)

    
    start = time.clock()
    # python krono2jira.py -s http://localhost:8080 -u mika.nokka@ambientia.fi -w kissa -q . -n kissa -p kissa
    Authenticate(JIRASERVICE,PSWD,USER)
    jira=DoJIRAStuff(USER,PSWD,JIRASERVICE)
    
    ###############################################################################
    # Testing name -> jira useraccount mapping theory
    #found example
    #user="nokka, mika"
    #account=CheckExcelName(jira,user)
    
    #not found example
    #user="noskka, mikfa"
    #account=CheckExcelName(jira,user)
    
    #if (account != "NOT EXIST"):
    #    logging.debug("User:{0} ---> Jira Account:{1}".format(user,account))
    #else:
    #    logging.debug("User:{0} ---> NO Jira account".format(user,account))
    
    
    #sys.exit(5)
    
    ###########END OF POC CODE ##################################################
    
    
    
    excel=excelfilepath+"/"+filename
    logging.debug ("Excel file:{0}".format(excel))

    Issues=defaultdict(dict) 

    wb= openpyxl.load_workbook(excel)
    #types=type(wb)
    #logging.debug ("Type:{0}".format(types))
    #sheets=wb.get_sheet_names()
    #logging.debug ("Sheets:{0}".format(sheets))
    CurrentSheet=wb[MainSheet] 
    
    

    
    print "=====>    Internal configuration:{0} , {1} ".format(ENV, ONCE)
 
    ##############################################################################################
    # Go through main excel sheet and find people names, ask Jira matching useraccount info
    # NOTE: Uses hardcoded sheet/column values
    # NOTE: As this handles first sheet, using used row/cell reading (buggy, works only for first sheet) 
    #
    i=DATASTARTSROW # brute force row indexing
    for row in CurrentSheet[('J{}:J{}'.format(DATASTARTSROW,CurrentSheet.max_row))]:  # go trough all column J  rows 
        for mycell in row:
            NAME=mycell.value
            logging.debug("ROW:{0} Name:{1}".format(i,mycell.value))
            Issues[NAME]={} # add to dictionary as master key (PERSON)
            
            #Just hardocoded operations, POC is one off
            
            THENAME=(CurrentSheet.cell(row=i, column=J).value)
            if not THENAME:
                THENAME="NO_NAME"
            Issues[NAME]["THENAME"] = THENAME
            
            account=CheckExcelName(jira,THENAME)
            Issues[NAME]["ACCOUNT"] = account
            CurrentSheet.cell(row=i, column=J).value=account
            
            logging.debug("---------------------------------------------------")
            i=i+1
    
    #print Issues.items() 
  
#>>> wb = Workbook()
         


    for key, value in Issues.iteritems() :
        logging.info("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        #print "The name: {0}".format(key)
        #print key, ">>>", value
        
        one=Issues.get(key)
        #for key, value in one.iteritems() :        
            #print key, "---->", value
            #if (key=="ACCOUNT" and value=="NOT EXIST"):
            #    logging.debug("ACCOUNTTIII EI OLE:{0}").format(THENAME)
        ACCOUNT=one.get('ACCOUNT') 
        #logging.debug("ACCOUNT:{0}".format(ACCOUNT))
        THENAME=one.get('THENAME')   
        #logging.debug("THENAME:{0}".format(THENAME))
        if (ACCOUNT=="NOT EXIST"):
            logging.info("{0} has NO Jira account".format(THENAME)) 
            
        else:
            logging.info("{0} --> Jira account: {1}".format(THENAME,ACCOUNT))      
            
            
                
        time.sleep(0.3) # prevent jira crashing for script attack
        if (ONCE=="YES"):
            print "ONCE testing mode ,stopping now"
            sys.exit(5) #testing do only once
        #print "++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        #now excel has been processed
    
    wb.save('WITH_ACCOUNT.xlsx')   
    end = time.clock()
    totaltime=end-start
    logging.info("\n*************************************************************************")
    logging.info("Time taken:{0} seconds".format(totaltime))
    logging.info("*************************************************************************")
    sys.exit(0)
    




 



##################################################################################33    
def CheckExcelName(jira,user):

   
    result=jira.search_users(user, startAt=0, maxResults=50, includeActive=True, includeInactive=False) #returns dictionary
    
    
    #logging.debug ("WHOLE STUFF:{0}".format(result))
    
    account ="NOT EXIST"
    # THIS WORKS
    logging.debug ("--------------------------------------------------------------------")
    
    if not result:
        logging.debug ("NO JIRA account MATCH for name:{0}".format(user))
        return (account)
    else:
        logging.debug ("Found matches for excel name:    {0}    --> checking now....".format(user))
    
    for user in result:
        #logging.debug ("DisplayName:{0}".format(user.displayName))
        #logging.debug ("Key:{0}".format(user.key))
    
        regex = r"(.*)(,)(.*)"   # Kehveli, Kalle  is username format in Jira
        match = re.search(regex, user.displayName)
                
        if (match):
            firstname=match.group(3).encode('utf-8')  # ääkköset off
            lastname=match.group(1).encode('utf-8')
            account=user.key
            logging.debug ("MATCH FOUND!!   Firstname:{0}   Secondname:{1}".format(firstname,lastname))
            logging.debug ("Has Jira user account: {0}".format(account))
            return account
        #else:
            #print "no match"
    logging.debug ("------------------------------------------------------------------")
    
    
    
    

    
if __name__ == '__main__':
    main()
    
    
    
    

    
    
    